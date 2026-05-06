import openpyxl
import sys

# Set encoding to utf-8 for Vietnamese characters
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

def inspect_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        print(f"Sheet: {sheet_name}")
        print(f"Headers: {headers}")
        if sheet.max_row > 1:
            data_row = [cell.value for cell in sheet[2]]
            print(f"First Data Row: {data_row}")
        print("-" * 20)

if __name__ == "__main__":
    inspect_excel("database/WarePro_Export_5-5-2026.xlsx")
